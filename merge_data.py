import tools.database as database
import xlrd
import os
import openpyxl

from tools.config import Profession, settings


def group_professions() -> list[list[Profession]]:
    """Объединяем в группу профессии, у которых одинаковые названия, но разные уровни
    Это нужно чтобы было проще добавлять нулевые профессии, т.к в описании нулевой профессии должны быть описания всех уровней"""
    professions = database.get_all_professions()

    group_profession_by_title: list[list[Profession]] = []
    for profession in professions:
        if profession.name in (group[0].name for group in group_profession_by_title):
            for group in group_profession_by_title:
                if profession.name == group[0].name: group.append(profession)
        else:
            group_profession_by_title.append([profession, ])
    return group_profession_by_title


def merge_professions(groups: list[list[Profession]]):
    """Нужно выделить среди всех одинаковых профессий разных уровней нулевую профессию.
    В описании и навыки которой засунуть сумму остальных одинаковых профессий
    Еще нужно изменить название добавив туда уровень разряда
    Пример: Профессия: Маляр, Уровень:3 -> Название: Маляр (3 разряда), Уровень: 3"""
    for group in groups:
        # Добавляем в группу нулевую профессию, содержащую сумму описаний и сумму требований
        zero_profession_title = group[0].name
        zero_profession_direction = group[0].direction
        zero_profession_descr = "|".join((proff.description for proff in group))
        zero_profession_skills = "|".join((proff.requirements for proff in group))

        database.add_profession(
            data=Profession(zero_profession_direction, zero_profession_title, 0, zero_profession_descr, zero_profession_skills), 
            table_name=settings.merged_table)

        for proff in group: # Меняем наименования профессий, добавляя к ним уровень разряда
            proff.name = f"{proff.name} ({proff.level} разряда)"
            database.add_profession(data=proff, table_name=settings.merged_table)


def find_edwica_professions_in_etks_db(edwica_file: str):
    """Метод принимает на вход excel-файл. Считывает всю колонку 'Наименование 2-го уровня' и по очереди ищет каждую профессию в базе
    спарсенных вакансий из ETKS.
    Если профессия найдена, то метод заберет из базы значение колонки 'description', поделит содержимое ячейки по разделителю | на разные функции
    и запишет их в отдельные колонки excel-файла на странице 'Функции (общие)', такой же принцип с колонкой БД 'skills', только теперь это значение
    запишется в отдельные колонки excel-файла на странице 'Навыки'
    Если профессия не найдена, то мы ее просто запишем в обе страницы 'Функции (общие)' и 'Навыки', чтобы соблюдать порядок"""
    if not edwica_file.endswith(".xlsx"): return "Файл должен быть формата .xlsx!"

    # Подключаем excel-файл
    book = xlrd.open_workbook(edwica_file)
    profession_sheet = book.sheet_by_name("Список профессий")
    wr_book = openpyxl.load_workbook(edwica_file)
    descr_sheet = wr_book['Функции (общие)']
    skill_sheet = wr_book['Навыки']

    # Собираем все профессии с эксель файла
    required_professions = (item.strip().lower() for item in profession_sheet.col_values(5)[1:] if item)
    # Подключаемся к БД с профессиями ETKS
    etks_professions = database.get_all_professions(table_name=settings.merged_table, only_zero_professions=True)
    for row, prof in enumerate(required_professions):
        # Инициализируем колонки на страницах "Функции" и "Навыки"
        descr_sheet.cell(row=row+2, column=1).value = row+1
        skill_sheet.cell(row=row+2, column=1).value = row+1
        # Заранее записываем профессию
        descr_sheet.cell(row=row+2, column=2).value = prof 
        skill_sheet.cell(row=row+2, column=2).value = prof 

        for etks_prof in etks_professions:
            if etks_prof.name.strip().lower() == prof:
                # Делим на составные части общее описание - каждая часть - это отдельный столбик и отдельный уровень
                descrs = etks_prof.description.split("|") 
                skills = etks_prof.requirements.split("|")
                for column, value in enumerate(descrs):
                    # Если нет заголовка колонки, то создаем ее
                    if not descr_sheet.cell(row=1, column=column+3).value: descr_sheet.cell(row=1, column=column+3).value = f"Функции {column+1}"
                    descr_sheet.cell(row=row+2, column=column+3).value = value
                for column, value in enumerate(skills):
                    if not skill_sheet.cell(row=1, column=column+3).value: skill_sheet.cell(row=1, column=column+3).value = f"Навык {column+1}"
                    skill_sheet.cell(row=row+2, column=column+3).value = value
                break 
    wr_book.save(edwica_file)

def main():
    import sys
    try:
        if sys.argv[1] == '--etks': # Группируем професии ETKS
            professions = group_professions()
            merge_professions(professions)
        
        elif sys.argv[1] == '--edwica': # Соединяем профессии ETKS с профессиями Edwica
            """Загружаем в папку data/excel/ необходимый список excel-файлов. Дальше программа берет каждый файл из папки и профессии из файла в БД"""
            for file in os.listdir(settings.excel_folder):
                if file.endswith('.xlsx'):
                    find_edwica_professions_in_etks_db(os.path.join(settings.excel_folder, file))
        else:
            raise SystemExit
    except (SystemExit, IndexError):
        exit('Используется флаги --etks ИЛИ --edwica\n--etks - Сгруппировать уровни ETKS\n--edwica - Найти в базе ETKS профессии из базы Эдвики и добавить в базу Эдвики информацию о недостающих профессиях')

if __name__ == "__main__":
    main()
